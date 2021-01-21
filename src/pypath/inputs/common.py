#!/usr/bin/env python
# -*- coding: utf-8 -*-

#
#  This file is part of the `pypath` python module
#
#  Copyright
#  2014-2021
#  EMBL, EMBL-EBI, Uniklinik RWTH Aachen, Heidelberg University
#
#  File author(s): Dénes Türei (turei.denes@gmail.com)
#                  Nicolàs Palacio
#                  Olga Ivanova
#
#  Distributed under the GPLv3 License.
#  See accompanying file LICENSE.txt or copy at
#      http://www.gnu.org/licenses/gpl-3.0.html
#
#  Website: http://pypath.omnipathdb.org/
#

from future.utils import iteritems
from past.builtins import xrange, range

import sys
import xlrd
if hasattr(xlrd, 'xlsx'):
    xlrd.xlsx.ensure_elementtree_imported(False, None)
    xlrd.xlsx.Element_has_iter = True
from xlrd.biffh import XLRDError
import pypath.share.session as session_mod
import pypath.share.common as common

_logger = session_mod.Logger(name = 'dataio')
_log = _logger._log
_console = _logger._console

try:
    import openpyxl
except:
    _log('No module `openpyxl` available.')

if 'unicode' not in __builtins__: unicode = str


def read_xls(
        xls_file,
        sheet = 0,
        use_openpyxl = False,
    ):
    """
    Generic function to read MS Excel XLS file, and convert one sheet
    to CSV, or return as a list of lists
    """

    table = []

    if not use_openpyxl:

        try:

            if hasattr(xls_file, 'read'):

                book = xlrd.open_workbook(
                    file_contents = xls_file.read(),
                    on_demand = True,
                )

            else:

                book = xlrd.open_workbook(xls_file, on_demand = True)

            try:
                if isinstance(sheet, int):
                    sheet = book.sheet_by_index(sheet)
                else:
                    sheet = book.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                sheet = book.sheet_by_index(0)

            table = [
                [common.basestring(c.value) for c in sheet.row(i)]
                for i in xrange(sheet.nrows)
            ]

            use_openpyxl = False

        except IOError:

            raise FileNotFoundError(xls_file)

        except:

            use_openpyxl = True

    if use_openpyxl and 'openpyxl' in sys.modules:

        try:

            book = openpyxl.load_workbook(
                filename = xls_file,
                read_only = True,
                on_demand = True,
            )

        except:

            raise ValueError('Could not open xls: %s' % xls_file)

            if not os.path.exists(xls_file):

                raise FileNotFoundError(xls_file)

        try:

            if type(sheet) is int:
                sheet = book.worksheets[sheet]
            else:
                sheet = book[sheet]

        except:

            sheet = book.worksheets[0]

        cells = sheet.get_squared_range(
            1, 1, sheet.max_column, sheet.max_row
        )

        table = [
            [common.basestring(c.value) if c.value else '' for row in cells]
            for c in row
        ]

    if 'book' in locals() and hasattr(book, 'release_resources'):

        book.release_resources()

    return table


def csv_sep_change(csv, old, new):

    clean_csv = []
    bw_quotes = False

    for char in csv:
        if char == '\r':
            continue
        elif char == '"':
            bw_quotes = not bw_quotes
        elif char == '\n':
            if not bw_quotes:
                clean_csv.append(char)
            else:
                clean_csv.append(' ')
        elif char == old:
            if bw_quotes:
                clean_csv.append(char)
            else:
                clean_csv.append(new)
        else:
            clean_csv.append(char)

    return ''.join(clean_csv)


def _try_isoform(name):

    name = name.split('-')

    if len(name) > 1 and name[1].isdigit():

        isoform = int(name[1])
        main = name[0]

    else:

        main = '-'.join(name)
        isoform = None

    return main, isoform


def read_table(
        cols,
        fileObject = None,
        data = None,
        sep = '\t',
        sep2 = None,
        rem = None,
        hdr = None,
        encoding = 'ascii',
    ):
    """
    Generic function to read data tables.

    fileObject : file-like
        Any file like object: file opened for read, or StringIO buffer
    cols : dict
        Dictionary of columns to read. Keys identifying fields are returned
        in the result. Values are column numbers.
    sepLevel1 : str
        Field separator of the file.
    sepLevel2 : dict
        Subfield separators and prefixes.
        E.g. {2: ',', 3: '|'}
    hdr : int
        Number of header lines. If None, no headers assumed.
    rem : list
        Strings to remove. For each line these elements will be replaced with ''.
    """

    rem = rem or []

    if data is None:

        if hasattr(fileObject, 'readline'):

            fileObject.seek(0)

        if hdr:

            for h in xrange(0, hdr):

                _ = next(fileObject)

        data = fileObject

    else:

        data = [l.strip() for l in data.split('\n') if len(l) > 0][hdr:]

    res = []

    for l in data:

        if type(l) is bytes:

            l = l.decode(encoding)

        for r in rem:

            l = l.replace(r, '')

        l = [f.strip() for f in l.split(sep)]

        if len(l) > max(cols.values()):

            dic = {}

            for name, col in iteritems(cols):

                field = l[col].strip()

                if sep2 is not None:

                    field = [
                        sf.strip()
                        for sf in field.split(sep2)
                        if len(sf) > 0
                    ]

                dic[name] = field

            res.append(dic)

    if fileObject is not None:

        fileObject.close()

    return res
